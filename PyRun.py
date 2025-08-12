
# =============================================
# 📌 임포트체커 (v250808 + 분석기능 + 중복표시 기능)
# 작성자: 서기대
# 주요 기능:
#   - 엑셀 기반 수용가 단말기 등록용 SQL VALUES 자동 생성
#   - 수용가 통계 분석 (수량, 항목 분류 등)
#   - 수용가번호 중복 항목 적색 표시 후 저장
# =============================================

import time
import os
import requests
import gspread
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from datetime import datetime
from tkinter import ttk
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# 📌 엑셀 컬럼명 정의 (20개 항목)
COLUMNS = [
    '수용가명', '수용가번호', '구주소', '신주소', '경도', '위도', '업종', '소속',
    '블록', '수용가 전화번호', '수용가 대상 년도', '검침원', '검침일', '계량기번호',
    '구경', '통신', '단말 부번호', '단말 주번호', '단말 회사', '단말 설치일'
]

SITE_SQ = 4
COMPANY_SQ = 9

def generate_sql_from_excel():
    file_path = filedialog.askopenfilename(title="엑셀 파일 선택", filetypes=[("Excel files", "*.xlsx *.xls")])
    if not file_path:
        messagebox.showwarning("파일 선택", "파일이 선택되지 않았습니다.")
        return

    idx = site_combobox.current()
    if idx < 0:
        messagebox.showwarning("계정명 선택", "먼저 계정명을 선택하세요.")
        return
    try:
        selected_num_len = int(filtered_df.iloc[idx]['수용가번호길이'])
    except Exception:
        messagebox.showerror("계정명 오류", "선택한 계정의 수용가번호길이 값이 올바르지 않습니다.")
        return

    try:
        df = pd.read_excel(file_path, dtype=str).fillna('')
        df = df.iloc[:, 1:1+len(COLUMNS)]
        df.columns = COLUMNS

        admin_no_counts = df['수용가번호'].value_counts()
        duplicated_admin_nos = set(admin_no_counts[admin_no_counts > 1].index)

        values_list = []
        success_count = 0
        fail_count = 0
        success_admin_no_list = []

        validation_cases = [
            ("수용가번호 길이 검사", lambda row: len(row['수용가번호']) == selected_num_len, "수용가번호 길이 불일치"),
            ("수용가번호 중복 검사", lambda row: row['수용가번호'] not in duplicated_admin_nos, "수용가번호 중복"),
            ("수용가 전화번호 길이 검사", lambda row: len(row['수용가 전화번호']) < 14, "수용가 전화번호 13자리 초과"),
        ]

        for idx_row, row in df.iterrows():
            try:
                for case_num, (desc, check_func, err_msg) in enumerate(validation_cases, 1):
                    if not check_func(row):
                        raise ValueError(f"[CASE #{case_num}] {err_msg} (값: {row['수용가번호']})")

                def safe_float(val, col):
                    if val == '':
                        raise ValueError(f"{col} 값이 비어있음")
                    return float(val)

                def safe_int(val, col):
                    if val == '':
                        raise ValueError(f"{col} 값이 비어있음")
                    return int(val)

                values = (
                    row['수용가명'], row['수용가번호'], row['구주소'], row['신주소'],
                    safe_float(row['경도'], '경도'), safe_float(row['위도'], '위도'),
                    row['업종'], SITE_SQ, COMPANY_SQ,
                    row['수용가 전화번호'], row['수용가 대상 년도'], row['검침원'],
                    safe_int(row['검침일'], '검침일'), row['계량기번호'],
                    safe_int(row['구경'], '구경'), row['통신'], row['단말 부번호'],
                    row['단말 주번호'], COMPANY_SQ,
                    f"{pd.to_datetime(row['단말 설치일']).date()}"
                )

                formatted = "('{}', '{}', '{}', '{}', {}, {}, '{}', {}, {}, '{}', '{}', '{}', {}, '{}', {}, '{}', '{}', '{}', {}, '{}'::timestamp)".format(*values)
                values_list.append(formatted)
                success_admin_no_list.append(row['수용가번호'])
                success_count += 1

            except Exception as e:
                values_list.append(f"-- [ERROR #{idx_row+1}] {e}")
                fail_count += 1

        result_text.config(state=tk.NORMAL)
        result_text.delete(1.0, tk.END)
        result_text.insert(tk.END, "-- ✅ 임포트전 조회할 수용가목록")
        result_text.insert(tk.END, ",".join(f"'{x}'" for x in success_admin_no_list) + "")
        result_text.insert(tk.END, f"-- 총 {len(df)}개 중 {success_count}개 성공, {fail_count}개 실패")
        result_text.insert(tk.END, ",".join(values_list))
        result_text.config(state=tk.DISABLED)

    except Exception as e:
        messagebox.showerror("에러 발생", str(e))

def analyze_excel_customer_stats():
    file_path = filedialog.askopenfilename(title="엑셀 파일 선택", filetypes=[("Excel files", "*.xlsx *.xls")])
    if not file_path:
        messagebox.showwarning("파일 선택", "파일이 선택되지 않았습니다.")
        return

    try:
        df = pd.read_excel(file_path, engine='openpyxl').fillna('')
        df.columns = [col.strip() for col in df.columns]

        result_text.config(state=tk.NORMAL)
        result_text.delete(1.0, tk.END)

        # 실제 컬럼명 출력
        result_text.insert(tk.END, "📋 엑셀 파일의 실제 컬럼명:\n")
        for i, col in enumerate(df.columns, 1):
            result_text.insert(tk.END, f"{i:2d}. {col}\n")
        result_text.insert(tk.END, "\n")

        # 수용가번호 총 수량
        if '수용가번호' in df.columns:
            total = df['수용가번호'].nunique()
            result_text.insert(tk.END, f"✅ 수용가번호 총 수량: {total}\n\n")
        else:
            result_text.insert(tk.END, "⚠️ '수용가번호' 컬럼이 없습니다.\n\n")

        # 일반적인 분석 컬럼들 (실제 컬럼명에 맞게 조정)
        analysis_columns = []
        
        # 블록 관련 컬럼 찾기
        block_columns = [col for col in df.columns if '블록' in col]
        analysis_columns.extend(block_columns)
        
        # 구분 관련 컬럼 찾기
        division_columns = [col for col in df.columns if '구분' in col or '분류' in col or '구' in col]
        analysis_columns.extend(division_columns)
        
        # 기타 분석 가능한 컬럼들
        other_columns = ['업종', '소속', '통신', '구경', '검침원']
        for col in other_columns:
            if col in df.columns:
                analysis_columns.append(col)
        
        # 중복 제거
        analysis_columns = list(set(analysis_columns))
        
        if analysis_columns:
            result_text.insert(tk.END, "📊 항목별 통계 분석:\n")
            for col in analysis_columns:
                if col in df.columns:
                    result_text.insert(tk.END, f"\n📈 '{col}' 항목별 개수:\n")
                    counts = df[col].value_counts()
                    total_count = len(counts)
                    result_text.insert(tk.END, f"총 {total_count}개 항목\n")
                    
                    # 상위 10개만 표시 (너무 많으면 화면이 복잡해짐)
                    display_counts = counts.head(10)
                    for value, count in display_counts.items():
                        percentage = (count / len(df)) * 100
                        result_text.insert(tk.END, f"- {value}: {count}개 ({percentage:.1f}%)\n")
                    
                    if len(counts) > 10:
                        result_text.insert(tk.END, f"... 외 {len(counts) - 10}개 항목\n")
                else:
                    result_text.insert(tk.END, f"⚠️ 컬럼 '{col}' 없음\n")
        else:
            result_text.insert(tk.END, "⚠️ 분석 가능한 컬럼을 찾을 수 없습니다.\n")

        result_text.config(state=tk.DISABLED)

    except Exception as e:
        messagebox.showerror("분석 오류", str(e))

def mark_duplicates_in_place():
    file_path = filedialog.askopenfilename(title="엑셀 파일 선택", filetypes=[("Excel files", "*.xlsx")])
    if not file_path:
        messagebox.showwarning("파일 선택", "파일이 선택되지 않았습니다.")
        return

    try:
        df = pd.read_excel(file_path, dtype=str).fillna('')
        if '수용가번호' not in df.columns:
            messagebox.showerror("열 없음", "'수용가번호' 열이 존재하지 않습니다.")
            return

        # 진행상태 창 생성
        progress_window = tk.Toplevel()
        progress_window.title("중복 검사 및 표시 진행상태")
        progress_window.geometry("500x200")
        progress_window.transient(window)
        progress_window.grab_set()

        # 프로그레스바
        progress_label = tk.Label(progress_window, text="중복 항목 검사 및 적색 표시 중...")
        progress_label.pack(pady=(20, 10))
        
        progress_bar = ttk.Progressbar(progress_window, length=400, mode='determinate')
        progress_bar.pack(pady=(0, 10))
        
        status_label = tk.Label(progress_window, text="")
        status_label.pack(pady=(0, 10))
        
        detail_label = tk.Label(progress_window, text="")
        detail_label.pack(pady=(0, 10))

        # 중복 검사할 항목들 정의 (지시부번호는 단말 주번호와 동일)
        duplicate_checks = [
            ('수용가번호', '수용가번호'),
            ('계량기번호', '계량기번호'),
            ('단말 주번호', '지시부번호/단말 주번호'),
            ('단말 부번호', '단말 부번호'),
            ('IMEI', 'IMEI'),
            ('패스워드', '패스워드')
        ]
        
        # 수용가상태 검사할 값들
        status_checks = ['단수', '중지', '철거', '폐전']
        
        # 자릿수 검사할 항목들 (컬럼명, 최소자릿수, 최대자릿수)
        digit_checks = [
            ('수용가번호', 13, 13),      # 수용가번호는 정확히 13자리
            ('계량기번호', 5, 20),      # 계량기번호는 보통 5-20자리
            ('단말 주번호', 5, 15),     # 단말 주번호는 보통 5-15자리
            ('IMEI', 15, 15),          # IMEI는 정확히 15자리
            ('패스워드', 4, 20)         # 패스워드는 보통 4-20자리
        ]
        
        # 실제 존재하는 컬럼만 필터링
        available_checks = []
        for col_name, display_name in duplicate_checks:
            if col_name in df.columns:
                available_checks.append((col_name, display_name))
        
        if not available_checks:
            messagebox.showerror("오류", "검사할 수 있는 컬럼이 없습니다.")
            return

        # 중복된 행들의 인덱스 수집
        duplicate_rows = set()
        duplicate_stats = {}
        duplicate_details = {}
        
        # 수용가상태 문제 행들의 인덱스 수집
        status_problem_rows = set()
        
        # 빈값/자릿수 문제 셀들의 위치 수집 (행, 열, 문제유형)
        cell_problems = []  # [(row_idx, col_name, problem_type, value), ...]
        cell_problem_stats = {}
        
        total_checks = len(available_checks)
        progress_bar['maximum'] = total_checks
        
        for check_idx, (col_name, display_name) in enumerate(available_checks):
            status_label.config(text=f"검사 중: {display_name}")
            detail_label.config(text=f"진행률: {check_idx + 1}/{total_checks}")
            progress_bar['value'] = check_idx + 1
            progress_window.update()
            
            # 중복 검사
            duplicated = df[col_name][df[col_name].duplicated(keep=False)]
            if not duplicated.empty:
                # 중복된 값들을 가진 행들의 인덱스 수집
                duplicated_indices = df[df[col_name].duplicated(keep=False)].index
                duplicate_rows.update(duplicated_indices)
                duplicate_stats[display_name] = len(duplicated)
                
                # 중복된 값들의 목록 저장 (상위 10개만)
                duplicated_values = duplicated.unique()
                duplicate_details[display_name] = duplicated_values[:10].tolist()
        
        # 수용가상태 검사
        if '수용가상태' in df.columns:
            status_label.config(text="수용가상태 검사 중...")
            detail_label.config(text="단수, 중지, 철거, 폐전 검사")
            progress_window.update()
            
            for status_value in status_checks:
                status_indices = df[df['수용가상태'] == status_value].index
                status_problem_rows.update(status_indices)
        
        # 빈값 및 자릿수 검사
        status_label.config(text="빈값 및 자릿수 검사 중...")
        detail_label.config(text="셀별 문제 검사")
        progress_window.update()
        
        # 빈값 검사
        empty_count = 0
        for col_name in ['수용가번호', '계량기번호', '단말 주번호', 'IMEI']:
            if col_name in df.columns:
                empty_indices = df[df[col_name] == ''].index
                for idx in empty_indices:
                    cell_problems.append((idx, col_name, '빈값', ''))
                    empty_count += 1
        
        if empty_count > 0:
            cell_problem_stats['빈값'] = empty_count
        
        # 자릿수 검사
        digit_count = 0
        for col_name, min_digits, max_digits in digit_checks:
            if col_name in df.columns:
                for idx, row in df.iterrows():
                    value = str(row[col_name])
                    if value and value != '':
                        if len(value) < min_digits or len(value) > max_digits:
                            cell_problems.append((idx, col_name, '자릿수', value))
                            digit_count += 1
        
        if digit_count > 0:
            cell_problem_stats['자릿수'] = digit_count
        
        # 모든 문제가 있는 행들 통합
        all_problem_rows = duplicate_rows.union(status_problem_rows)
        
        if not all_problem_rows:
            messagebox.showinfo("문제 없음", "중복된 항목이나 문제가 있는 수용가상태가 없습니다.")
            progress_window.destroy()
            return

        # 엑셀 파일에 색상 표시
        status_label.config(text="엑셀 파일에 색상 표시 중...")
        detail_label.config(text=f"총 {len(all_problem_rows)}개 행 처리")
        progress_window.update()
        
        wb = load_workbook(file_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        
        # 색상 및 음영 정의
        red_font = Font(color="FF0000")      # 적색: 수용가상태 문제
        blue_font = Font(color="0000FF")     # 파란색: 중복 값
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 노란색 음영
        
        # 수용가상태 문제 행들을 적색으로 표시 (행 전체)
        for row_idx in status_problem_rows:
            excel_row = row_idx + 2  # pandas는 0-based, excel은 1-based + 헤더
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=excel_row, column=col)
                cell.font = red_font
        
        # 중복된 값이 있는 셀들을 노란색 음영으로 표시 (셀별)
        for col_name, display_name in available_checks:
            if col_name in df.columns:
                duplicated_values = df[col_name][df[col_name].duplicated(keep=False)]
                if not duplicated_values.empty:
                    duplicated_indices = df[df[col_name].duplicated(keep=False)].index
                    
                    # 헤더에서 해당 컬럼의 위치 찾기
                    try:
                        col_idx = headers.index(col_name) + 1
                        # 중복된 값이 있는 셀들만 노란색 음영으로 표시
                        for row_idx in duplicated_indices:
                            excel_row = row_idx + 2  # pandas는 0-based, excel은 1-based + 헤더
                            cell = ws.cell(row=excel_row, column=col_idx)
                            cell.fill = yellow_fill
                    except ValueError:
                        # 컬럼이 존재하지 않는 경우 무시
                        pass
        
        # 빈값/자릿수 문제 셀들을 노란색 음영으로 표시
        for row_idx, col_name, problem_type, value in cell_problems:
            excel_row = row_idx + 2  # pandas는 0-based, excel은 1-based + 헤더
            
            # 헤더에서 해당 컬럼의 위치 찾기
            try:
                col_idx = headers.index(col_name) + 1
                cell = ws.cell(row=excel_row, column=col_idx)
                cell.fill = yellow_fill
            except ValueError:
                # 컬럼이 존재하지 않는 경우 무시
                pass
        
        wb.save(file_path)
        progress_window.destroy()
        
        # 통계 표시
        stats_text = "🔍 검사 결과:\n\n"
        
        # 중복 통계
        if duplicate_stats:
            stats_text += "📊 중복 항목:\n"
            for display_name, count in duplicate_stats.items():
                stats_text += f"  🔵 중복 {display_name}: {count}개\n"
                if display_name in duplicate_details:
                    details = duplicate_details[display_name]
                    if len(details) > 0:
                        stats_text += "     중복 값 예시: " + ", ".join(str(x) for x in details[:5])
                        if len(details) > 5:
                            stats_text += f" ... 외 {len(details) - 5}개"
                        stats_text += "\n"
        
        # 수용가상태 통계
        if status_problem_rows:
            status_counts = {}
            for status_value in status_checks:
                if '수용가상태' in df.columns:
                    count = len(df[df['수용가상태'] == status_value])
                    if count > 0:
                        status_counts[status_value] = count
            
            if status_counts:
                stats_text += "\n📊 수용가상태 문제:\n"
                for status_value, count in status_counts.items():
                    stats_text += f"  🔴 {status_value}: {count}개\n"
        
        # 셀별 문제 통계
        if cell_problem_stats:
            stats_text += "\n📊 셀별 문제:\n"
            for problem_type, count in cell_problem_stats.items():
                stats_text += f"  🟡 {problem_type}: {count}개\n"
        
        stats_text += f"\n총 {len(all_problem_rows)}개 행이 색상으로 표시되었습니다."
        stats_text += f"\n🟡 노란색 음영: 중복 항목 ({len(duplicate_rows)}개)"
        stats_text += f"\n🔴 적색: 수용가상태 문제 ({len(status_problem_rows)}개)"
        stats_text += f"\n🟡 노란색 음영: 빈값/자릿수 문제 ({len(cell_problems)}개)"
        
                # 시트 추가 옵션 제공
        response = messagebox.askyesno("검사 완료", 
                                      f"{stats_text}\n\n"
                                      f"원본 파일에 '중복항목' 시트를 추가하시겠습니까?")
        
        if response:
            create_filtered_file(file_path, list(all_problem_rows), df)

    except Exception as e:
        messagebox.showerror("오류", str(e))

def create_filtered_file(original_file_path, duplicate_row_indices, original_df):
    """원본 파일에 중복 항목 시트 추가"""
    try:
        # 중복된 행들만 필터링
        filtered_df = original_df.iloc[duplicate_row_indices].copy()
        
        if filtered_df.empty:
            messagebox.showinfo("결과", "중복된 항목이 없습니다.")
            return
        
        # 원본 파일 열기
        wb = load_workbook(original_file_path)
        
        # 기존에 '중복항목' 시트가 있다면 삭제
        if '중복항목' in wb.sheetnames:
            wb.remove(wb['중복항목'])
        
        # 새 시트 생성
        ws_new = wb.create_sheet('중복항목')
        
        # 헤더 복사 (첫 번째 시트에서)
        ws_original = wb.active
        headers = []
        for col in range(1, ws_original.max_column + 1):
            header_value = ws_original.cell(row=1, column=col).value
            headers.append(header_value)
            ws_new.cell(row=1, column=col, value=header_value)
        
        # 데이터 복사
        for row_idx, original_row_idx in enumerate(duplicate_row_indices, 2):
            excel_row = original_row_idx + 2  # pandas는 0-based, excel은 1-based + 헤더
            for col in range(1, ws_original.max_column + 1):
                cell_value = ws_original.cell(row=excel_row, column=col).value
                ws_new.cell(row=row_idx, column=col, value=cell_value)
        
        # 중복된 행들 전체를 적색으로 표시
        red_font = Font(color="FF0000")
        
        # 모든 행을 적색으로 표시 (행 전체)
        for row in range(2, ws_new.max_row + 1):
            for col in range(1, ws_new.max_column + 1):
                cell = ws_new.cell(row=row, column=col)
                cell.font = red_font
        
        # 파일 저장
        wb.save(original_file_path)
        
        messagebox.showinfo("완료", f"원본 파일에 '중복항목' 시트가 추가되었습니다.\n총 {len(duplicate_row_indices)}개 행이 포함되었습니다.")
        
    except Exception as e:
        messagebox.showerror("시트 추가 오류", str(e))

def read_google_sheet(sheet_url, credentials_path, worksheet_name):
    gc = gspread.service_account(filename=credentials_path)
    sh = gc.open_by_url(sheet_url)
    worksheet = sh.worksheet(worksheet_name)
    all_values = worksheet.get_all_values()
    header = all_values[0]
    idx_account = header.index('계정명')
    idx_service = header.index('서비스코드')
    idx_len = header.index('수용가번호길이')
    idx_struct = header.index('고객번호구조')
    data = [
        [row[idx_account], row[idx_service], row[idx_len], row[idx_struct]]
        for row in all_values[1:] if row[idx_account] and row[idx_service]
    ]
    df = pd.DataFrame(data, columns=['계정명', '서비스코드', '수용가번호길이', '고객번호구조'])
    return df

# GUI 구성
sheet_url = "https://docs.google.com/spreadsheets/d/10XO7o99fYr4e_I_etJF3_eCFa2_dsDs2egSWeu1GAls/edit#gid=679649875"
credentials_path = r"C:\제품등록\gcp9304-4410543fedf2.json"
worksheet_name = "IN형식"
df = read_google_sheet(sheet_url, credentials_path, worksheet_name)

window = tk.Tk()
window.title("임포트체커 + 통계 + 중복표시 (v250701)")
window.geometry("1000x600")

# 녹색 램프 프레임 (왼쪽 상단)
lamp_frame = tk.Frame(window)
lamp_frame.pack(anchor=tk.NW, padx=10, pady=(10, 0))

# 녹색 램프 (작은 네모박스)
lamp_canvas = tk.Canvas(lamp_frame, width=20, height=10, bg='white', highlightthickness=1, highlightbackground='black')
lamp_canvas.pack(side=tk.LEFT)

# 램프 상태 변수
lamp_on = True

def toggle_lamp():
    global lamp_on
    if lamp_on:
        lamp_canvas.configure(bg='lightgreen')
    else:
        lamp_canvas.configure(bg='white')
    lamp_on = not lamp_on
    # 1초마다 반복
    window.after(1000, toggle_lamp)

# 램프 시작
toggle_lamp()

site_frame = tk.Frame(window)
site_frame.pack(fill=tk.X, padx=10, pady=(10, 0))

site_label = tk.Label(site_frame, text="계정명 선택:")
site_label.pack(side=tk.LEFT)

exclude_accounts = ['나라장터', '농촌공사', '로우리스', '']
filtered_df = df[~df['계정명'].isin(exclude_accounts)].copy()
filtered_df = filtered_df.sort_values('계정명').reset_index(drop=True)
site_combobox = ttk.Combobox(site_frame, values=list(filtered_df['계정명']), state="readonly")
site_combobox.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 0))

def on_site_select(event):
    idx = site_combobox.current()
    if idx >= 0:
        account_name = filtered_df.iloc[idx]['계정명']
        service_code = filtered_df.iloc[idx]['서비스코드']
        num_len = filtered_df.iloc[idx]['수용가번호길이']
        struct = filtered_df.iloc[idx]['고객번호구조']
        messagebox.showinfo("선택한 계정", f"계정명: {account_name}\n서비스코드: {service_code}\n수용가번호길이: {num_len}\n고객번호구조: {struct}")
site_combobox.bind('<<ComboboxSelected>>', on_site_select)

btn_sql = tk.Button(window, text="엑셀 파일 선택 및 SQL 변환", command=generate_sql_from_excel, bg="lightblue")
btn_sql.pack(pady=(10, 5))

btn_analyze = tk.Button(window, text="엑셀 파일 선택 및 수용가 통계 분석", command=analyze_excel_customer_stats, bg="lightgreen")
btn_analyze.pack(pady=(0, 5))

btn_dup = tk.Button(window, text="🔍 종합검사 (적색/노란색 음영 표시)", command=mark_duplicates_in_place, bg="salmon")
btn_dup.pack(pady=(0, 10))

frame = tk.Frame(window)
frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

x_scrollbar = tk.Scrollbar(frame, orient=tk.HORIZONTAL)
x_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)

y_scrollbar = tk.Scrollbar(frame)
y_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

result_text = tk.Text(frame, wrap=tk.NONE, xscrollcommand=x_scrollbar.set, yscrollcommand=y_scrollbar.set)
result_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

x_scrollbar.config(command=result_text.xview)
y_scrollbar.config(command=result_text.yview)

window.mainloop()
